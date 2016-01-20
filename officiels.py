# -* -coding: utf-8 -*-

"""
.. module: models.py
   :platform: Unix, Windows
   :synopsys: List of the models for tickets structure

.. moduleauthor: Cedric Airaud <cairaud@gmail.com>
"""

import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import requests
import argparse
import datetime

import logging
logging.basicConfig(level=logging.INFO, format="%(levelname)-9s %(lineno)-4s %(message)s")

import gen_pdf

jury_url = "http://ffn.extranat.fr/webffn/resultats.php?idact=nat&idcpt={competition}&go=off"
clubs_url = "http://ffn.extranat.fr/webffn/resultats.php?idact=nat&idcpt={competition}&go=clb"


class CompetitionException (Exception):
    pass

class ReunionException (Exception):
    pass

class Competition:
    """
    Represent a competition, composed of several Reunions
    """
    NIVEAU_DEPARTEMENTAL = "Départemental"
    NIVEAU_REGIONAL = "Régional"

    def __init__(self, conf, competition_index):
        """
        :param conf Configuration structure
        :type conf Configuration
        :param competition_index: Index of the competition
        :type competition_index int
        """
        self.index = competition_index

        # Get some info from configuration
        self.date = conf.competitions[competition_index]["date"]
        self.equipe = conf.competitions[competition_index]["equipe"]
        self.niveau = conf.competitions[competition_index]["niveau"]

        if self.equipe:
            try:
                self.equipe = int(self.equipe)
            except ValueError as e:
                logging.fatal("La colonne Equipe doit être un nombre pour une compétition par équipe")

        url = jury_url.format(competition=competition_index)
        logging.debug("Jury et réunions: " + url)
        data = requests.get(url).text
        soup = BeautifulSoup(data, 'html.parser')

        entete = soup.find("fieldset", class_="enteteCompetition")
        spans = entete.find_all("span")
        self.titre, self.type = spans[0].text, entete.text.splitlines()[-1]
        logging.info("{} - {} - {} ".format(self.type, self.titre, self.date.strftime("%d/%m/%Y")))

        reunions = []
        self.reunions = []
        self.participations = {}
        table = entete.find_next_sibling("table")

        try:
            for tr in table.find_all("tr"):
                tds = tr.find_all("td")
                if tds[0]['id'] == "mainResEpr":
                    reunion = Reunion(self, titre=tds[0].text.strip(), index=len(reunions))
                    reunions.append(reunion)
                    logging.debug("Réunion trouvée: " + str(reunion))
                else:
                    if len(tds) != 3:
                        logging.fatal("Besoin de 3 colonnes par officiel: " + tds.text)
                    if not reunion:
                        logging.fatal("Pas d'entête de réunion trouvé: " + tds.text)
                    poste, nom, club = tds[0].text.replace(":", "").strip(), tds[1].text, tds[2].text
                    if poste in conf.postes and not conf.postes[poste]:
                        logging.debug("{} au poste {} est ignoré".format(nom, poste))
                    elif club in conf.clubs:
                        officiel = conf.find_officiel(nom=nom, club=club)
                        logging.debug("Officiel trouvé: " + str(officiel))
                        if officiel not in reunion.officiels and conf.check_poste(officiel, poste):
                            reunion.officiels.append(officiel)
                    elif club != "NATATION AZUR":
                        logging.warning("Officiel ignoré: {} car le club {} n'est pas dans la liste".format(nom, club))

            # Not enough officiels for a reunion: ignore it
            for reunion in reunions:
                if len(reunion.officiels) < 5:
                    logging.warning("La réunion {} est ignorée: {} officiels".format(reunion.titre,
                                                                                     len(reunion.officiels)))
                else:
                    self.reunions.append(reunion)

            # Parse participations
            url = clubs_url.format(competition=competition_index)
            data = requests.get(url).text

            soup = BeautifulSoup(data, 'html.parser')

            table = soup.find("table", class_="tableau")
            for tr in table.find_all("tr"):
                tds = tr.find_all("td")
                if len(tds) == 13:
                    tds[1].b.clear()
                    club, num = tds[1].a.text.strip(), int(tds[4].text)
                    if club in conf.clubs:
                        self.participations[club] = num
                        conf.clubs[club].competitions.append(self)
                    else:
                        logging.warning("Club {} ignoré pour les participations car pas dans la liste".format(club))

        except KeyError as e:
            logging.warning("Pas de résultats pour la compétition {} du {}".format(self.titre,
                                                                                   self.date.strftime("%D/%m/%Y")))

    def __str__(self):
        return "{titre}\n{type}\n{date}\n\n".format(**self.__dict__) + "\n\n".join(map(str, self.reunions))

    def link(self):
        return "C{}".format(self.index)


class Reunion:
    """
    Represent a Reunion, base for the calculation
    """

    def __init__(self, competition, titre, index):
        self.competition = competition
        self.titre = titre
        self.index = index
        self.officiels = []
        self._officiels_per_club = None
        self.pts = {}
        self.details = {}

    def __str__(self):
        return self.titre + "\n  " + "\n  ".join(map(str, self.officiels))

    def officiels_per_club(self):
        """
        Sort officiels per club
        """
        if self._officiels_per_club:
            return self._officiels_per_club

        self._officiels_per_club = {}
        for officiel in self.officiels:
            if officiel.club.nom not in self._officiels_per_club:
                self._officiels_per_club[officiel.club.nom] = []
            self._officiels_per_club[officiel.club.nom].append(officiel)

        return self._officiels_per_club

    def points(self, club, details=None):
        """
        :param club: Club to look for
        :type club: Club
        :param details: Optional list to get detail about calculation
        :type details: None|List
        :return: Number of points
        :rtype: int
        """
        if club in self.pts and (club in self.details or details is None):
            if details is not None:
                details += self.details[club]
            return self.pts[club]

        participations = self.competition.participations.get(club.nom, 0)

        # needed = (Num of A/B, Total num)
        if self.competition.equipe:
            participations = participations // self.competition.equipe
            if participations <= 1:
                needed = (participations, participations)
            else:
                needed = (1, min(3, participations))

        else:
            if participations <= 1:
                needed = (0, 0)
            elif participations < 10:
                needed = (0, 1)
            elif participations < 20 or competition.niveau == Competition.NIVEAU_REGIONAL:
                needed = (1, 2)
            else:
                needed = (1, 3)

        if type(details) is list:
            s = "{} officiels requis".format(needed[1])
            if needed[0] > 0:
                s += ", dont {} A ou B".format(needed[0])
            details.append(s)

        num_ab, num = 0, 0
        club_officiels = self.officiels_per_club().get(club.nom, [])
        for officiel in club_officiels:
            num += 1
            level = officiel.get_level(self.competition.date)
            if level == 'A' or level == 'B':
                num_ab += 1

        if competition.niveau == Competition.NIVEAU_REGIONAL and num > 5:
            if type(details) is list:
                details.append("5 officiels retenus sur les {} présentés".format(num))
            num = 5

        # TODO: Officiel manquant -4 point, sans bon statut -2 points. Dans quel ordre ?
        if num < needed[1]:
            missing = needed[1] - num
            pts = missing * -4
            if type(details) is list:
                details.append("{} points négatifs pour {} officiels manquants".format(-pts, missing))
        else:
            extra = num - needed[1]
            pts = extra * 2
            if extra > 0 and type(details) is list:
                details.append("{} points supplémentaires pour {} officiels".format(pts, extra))
            if num_ab < needed[0]:
                missing = needed[0] - num_ab
                pts += missing * -2
                if type(details) is list:
                    details.append("{} points de malus par manque d'officiel A/B".format(missing*2))

        if details is not None:
            self.details[club] = details
        self.pts[club] = pts
        return pts

    def link(self):
        return "{}_R{}".format(self.competition.index, self.index)


class Officiel:
    """
    Represent an Officiel
    """
    def __init__(self, nom, club, a_depuis=None, b_depuis=None, c_depuis=None, valide=None):
        self.nom = nom
        self.club = club
        self.a_depuis = a_depuis
        self.b_depuis = b_depuis
        self.c_depuis = c_depuis
        self.valide = valide
        self.index = 0

    def __str__(self):
        return "{nom} ({club})".format(**self.__dict__)

    def get_level(self, date):
        """
        Return the level for an officiel as a string, at the given date (if specified)
        """
        if self.a_depuis and self.a_depuis < date:
            return 'A'
        elif self.b_depuis and self.b_depuis < date:
            return 'B'
        else:
            return 'C'


class Club:
    """
    Club
    """
    def __init__(self, nom, departement, index):
        self.nom = nom
        self.departement = departement
        self.index = index
        self.competitions = []

    def __str__(self):
        return "{} ({})".format(self.nom, self.departement)

    def link(self):
        return "Club{}".format(self.index)


class Configuration:
    """
    Global configuration
    """
    def __init__(self, filename):
        self.officiels = {}
        self.clubs = {}
        self.postes = {}
        self.competitions = {}
        self.dirty = False
        self.filename = filename

        self.wb = load_workbook(filename, guess_types=True)
        logging.info("Configuration depuis le fichier '{}'".format(filename))

        self.sheets = {'clubs': 'Clubs', 'officiels': 'Officiels', 'postes': 'Postes', 'competitions': 'Compétitions'}
        if len(set(self.wb.get_sheet_names()) & set(self.sheets.values())) != 4:
            raise Exception("Le fichier {} doit contenir les pages {} (Trouvées: {})".format(
                filename, ', '.join(self.sheets.values()), ', '.join(self.wb.get_sheet_names())))

        logging.info("- Lecture des clubs")
        xl_sheet = self.wb.get_sheet_by_name(self.sheets['clubs'])
        row = xl_sheet.rows[0]
        if row[0].value != "Club" or row[1].value != "Département":
            raise Exception("La page 'Clubs' doit contenir des colonnes 'Club' et 'Département' (Trouvées: {})".format(
                ", ".join([cell.value for cell in row])))
        for num, row in enumerate(xl_sheet.rows[1:]):
            if row[0].value != "":
                club = Club(nom=row[0].value, departement=row[1].value, index=num+1)
                self.clubs[club.nom] = club

        logging.info("- Lecture des officiels")
        xl_sheet = self.wb.get_sheet_by_name(self.sheets['officiels'])
        row = xl_sheet.rows[0]
        if (row[0].value != "Nom" or row[1].value != "Club" or row[2].value != "A depuis" or
            row[3].value != "B depuis" or row[4].value != "C depuis" or row[5].value != "Valide"):
            raise Exception("La page 'Officiels' doit contenir des colonnes 'Nom', 'Club', 'A depuis', 'B depuis' "
                            "'C depuis' et 'Valide' (Trouvées: {})".format(", ".join([cell.value for cell in row])))
        for index, row in enumerate(xl_sheet.rows[1:]):
            if row[0].value != "":
                club = row[1].value
                if club not in self.clubs:
                    print("WARNING: Le club {} pour l'officiel {} n'a pas été trouvé".format(club, row[0].value))
                else:
                    club = self.clubs[club]
                    officiel = Officiel(nom=row[0].value, club=club, a_depuis=row[2].value, b_depuis=row[3].value,
                                        c_depuis=row[4].value, valide=row[5].strip().lower() == "Oui")
                    officiel.index = index
                    self.officiels[officiel.nom] = officiel

        logging.info("- Lecture des postes")
        xl_sheet = self.wb.get_sheet_by_name(self.sheets['postes'])
        row = xl_sheet.rows[0]
        if row[0].value != "Poste" or row[1].value != "Niveau":
            raise Exception("La page 'Postes' doit contenir des colonnes 'Postes' et 'Niveau' "
                            "(Trouvées: {})".format(", ".join([cell.value for cell in row])))
        for row in xl_sheet.rows[1:]:
            if row[0].value != "":
                self.postes[row[0].value] = row[1].value

        logging.info("- Lecture des compétitions")
        xl_sheet = self.wb.get_sheet_by_name(self.sheets["competitions"])
        row = xl_sheet.rows[0]
        if (row[0].value != "Numéro" or row[1].value != "Date" or row[2].value != "Compétition" or
                row[3].value != "Niveau" or row[4].value != "Équipe"):
            raise Exception("La page 'Compétition' doit contenir des colonnes 'Numéro', 'Date' "
                            "'Compétition', 'Niveau' et 'Équipe' "
                            "(Trouvées: {})".format(", ".join([cell.value for cell in row])))
        for row in xl_sheet.rows[1:]:
            if row[0].value:
                self.competitions[row[0].value] = {"date": row[1].value,
                                                   "titre": row[2].value,
                                                   "niveau": row[3].value,
                                                   "equipe": row[4].value}

    def find_officiel(self, nom, club):
        """
        Find an officiel by name if it exists
        """
        if nom not in self.officiels:
            logging.warning("L'officiel {} (Club {}) n'existe pas".format(nom, club))
            officiel = Officiel(nom, self.clubs[club])
            self.officiels[nom] = officiel
            sheet = self.wb.get_sheet_by_name(self.sheets['officiels'])
            num_rows = len(sheet.rows)
            sheet.cell(row=num_rows+1, column=1, value=nom)
            sheet.cell(row=num_rows+1, column=2, value=club)
            officiel.index = num_rows
            self.dirty = True

        return self.officiels[nom]

    def update_officiel(self, officiel):
        """
        Update an officiel in the file
        :param officiel: Officiel to update
        :type officiel: Officiel
        """
        sheet = self.wb.get_sheet_by_name(self.sheets['officiels'])
        sheet.cell(row=officiel.index+1, column=1, value=officiel.nom)
        sheet.cell(row=officiel.index+1, column=2, value=officiel.club.nom)
        sheet.cell(row=officiel.index+1, column=3, value=officiel.a_depuis)
        sheet.cell(row=officiel.index+1, column=4, value=officiel.b_depuis)
        sheet.cell(row=officiel.index+1, column=5, value=officiel.c_depuis)
        sheet.cell(row=officiel.index+1, column=6, value="Oui" if officiel.valide else "Non")
        self.dirty = True


    def check_poste(self, officiel, poste):
        """
        Check that the poste matches the level for the Officiel
        :param officiel Officiel to check
        :type officiel Officiel
        :param poste Name of the poste
        :type poste basestring
        """
        if poste not in self.postes:
            logging.error("Le poste '{}' n'est pas listé dans le fichier de configuration".format(poste))
            return False

        niveau = conf.postes[poste]
        if niveau == 'A' and not officiel.a_depuis:
            logging.error("L'officiel {} semble avoir le niveau A".format(officiel.nom))
            return False

        elif niveau == 'B' and not officiel.b_depuis:
            logging.error("L'officiel {} semble avoir le niveau B".format(officiel.nom))
            return False

        return True

    def save(self):
        """
        Save the file if it has been updated
        """
        if self.dirty:
            backup_filename = self.filename + ".bak"
            logging.info("Mise à jour du fichier {} (Sauvegarde: {})".format(self.filename, backup_filename))
            os.rename(self.filename, backup_filename)
            try:
                self.wb.save(self.filename)
            except Exception as e:
                os.rename(backup_filename, self.filename)
                logging.error("Erreur lors de la mise à jour, restoration de la sauvegarde.\n" + str(e))
            self.dirty = False


    def update_officiels(self, filename):
        """
        Update the list of officiels
        :param filename: File to load
        :type filename: string
        """
        wb = load_workbook(args.update, guess_types=True)
        sheet = wb.active
        logging.info("Mise à jour des officiels depuis le fichier {}".format(filename))

        header = [cell.value for cell in sheet.rows[0]]
        labels = {}
        for label in ('nom', 'prenom', 'niveau_libelle', 'date_obtention_fr', 'validite_periode', 'dernier_club'):
            try:
                labels[label] = header.index(label)
            except ValueError:
                logging.fatal("Pas de cellule '{}' trouvée".format(label))

        for row in sheet.rows[1:]:
            if row[labels['validite_periode']].value != 'titre_perime':
                name = row[labels['nom']].value + " " + row[labels['prenom']].value
                date = datetime.datetime.strptime(row[labels['date_obtention_fr']].value, "%d/%m/%Y")
                club = row[labels['dernier_club']].value
                level = row[labels['niveau_libelle']].value

                if club in self.clubs:
                    officiel = self.find_officiel(name, club)
                    if level in ("Officiel A") and not officiel.a_depuis:
                        logging.info("Mise à jour du niveau A pour {}".format(str(officiel)))
                        officiel.a_depuis = date
                        self.update_officiel(officiel)
                    if level in ("Officiel A", "Officiel B", "Officiel A stagiaire") and not officiel.b_depuis:
                        officiel.b_depuis = date
                        logging.info("Mise à jour du niveau B pour {}".format(str(officiel)))
                        self.update_officiel(officiel)
                    if level in ("Officiel A", "Officiel B", "Officiel A stagiaire", "Officiel B stagiaire",
                                 "Officiel C") and not officiel.c_depuis:
                        officiel.c_depuis = date
                        logging.info("Mise à jour du niveau B pour {}".format(str(officiel)))
                        self.update_officiel(officiel)
                else:
                    logging.warning("L'officiel {} est ignoré (Club {})".format(name, club))

            else:
                logging.debug("Licence périmée pour l'officiel {}".format(name))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Liste des compétitions')
    parser.add_argument("--conf", default="Officiels.xlsx", help="Fichier de configuration")
    parser.add_argument("--update", help="Mettre à jour la liste des officiels")

    args = parser.parse_args()

    conf = Configuration('Officiels.xlsx')
    if args.update:
        conf.update_officiels(args.update)
        conf.save()
        exit()

    competitions = []

    doc = gen_pdf.DocTemplate(conf, "Compétitions.pdf", "Liste des compétitions", "Cédric Airaud")
    for competition_index in sorted(conf.competitions.keys()):
        competitions.append(Competition(conf, competition_index))
        conf.save()

    points = {"Départemental": {"participations": 0},
              "Régional":      {"participations": 0}}

    for competition in competitions:
        if competition.niveau in points:
            l = points[competition.niveau]
        else:
            l = points["Régional"]

        for club_name, num in competition.participations.items():
            l["participations"] += num
            if club_name not in l:
                    l[club_name] = 0;

        for reunion in competition.reunions:
            for club_name in competition.participations.keys():
                club = conf.clubs[club_name]
                pts = reunion.points(club, details=[])
                l[club_name] += pts

    doc.participations = {level: l["participations"] for level, l in points.items()}

    for club in sorted(conf.clubs.values(), key=lambda x: "{} {}".format(x.departement, x.nom)):
        doc.new_club(club)

    for competition in competitions:
        doc.new_competition(competition)

    doc.build()












